from tkinter import *
from tkinter import ttk
import openpyxl
import datetime

root = Tk()

root.title("Rent Calculator")
mainframe = ttk.Frame(root, padding = "3 3 12 12")
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

#print("Enter the name of the file")
file_name = StringVar()
status = StringVar()
rent_month = StringVar()
rent_year = StringVar()



sheet_name = "RENT CALCS"



#Formula to get the last day of the month
def last_day_of_month(any_day):
    next_month = any_day.replace(day = 28)+ datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)


#Formula to see if the container checkout date comes before the first of the month
def compare_dates(start, end):
    if start <= end:
        return True
    else:
        return False
#Sample code that prints every cell value in whatever range you enter
""" for row in range(2, ws.max_row+1):
    for column in "ABCDEF":
        cell_name = "{}{}".format(column, row)
        print(ws[cell_name].value) """

#All of this is just to parse the date from a string to the datetime format
def calculate_rent(*args):
    rent_month_string = int(rent_month.get())
    rent_year_string = int(rent_year.get())
    file_name_string = str(file_name.get())
    print(rent_year_string, rent_month_string, file_name_string)
    start_of_the_month = datetime.datetime(rent_year_string, rent_month_string, 1)
    before_month_start = start_of_the_month - datetime.timedelta(days=1)
    end_of_the_month = last_day_of_month(start_of_the_month)
    #for simplicity we're just going to have one spreadsheet with one page.
    print(file_name_string)
    wb = openpyxl.load_workbook(file_name_string)
    ws = wb[sheet_name]
    for row in range(2, ws.max_row+1):
        start_date = datetime.date(2020,1,1)
        end_date = datetime.date(2020,1,1)
        for column in "E":
            cell_name = "{}{}".format(column, row)
            if ws[cell_name].value == None:
                #print(ws[cell_name].value)
                start_date = start_of_the_month
            elif compare_dates(str(ws[cell_name].value), str(start_of_the_month)):
                start_date = before_month_start
            elif type(ws[cell_name].value) is str:
                print(ws[cell_name])
                start_date = datetime.datetime.strptime(ws[cell_name].value, '%Y-%b-%d')
            else:
                start_date = ws[cell_name].value
        for column in "F":
            cell_name = "{}{}".format(column, row)
            if ws[cell_name].value == None:
                #print(ws[cell_name].value)
                end_date = end_of_the_month
            elif type(ws[cell_name].value) is str:
                print(ws[cell_name].row)
                end_date = datetime.datetime.strptime(ws[cell_name].value, '%Y-%b-%d')
            else:
                end_date = ws[cell_name].value
        for column in "G":
            cell_name = "{}{}".format(column, row)
            if type(end_date) is str:
                end_date = datetime.datetime.strptime(end_date, '%Y-%b-%d')
            if type(start_date) is str:
                start_date = datetime.datetime.strptime(start_date, '%Y-%b-%d')        
            #Important note: should this be left as is or +1? Difference is leaving off one day, which will make a difference in the way rent is caluclated
            difference = (end_date - start_date).days 
            ws[cell_name] = difference
    #Gotta save before you read the entries you just wrote
    wb.save(filename=file_name_string)


    #This is to calculate the percent of rent due.
    for row in range(2, ws.max_row+1):
        rent_percentage = 0
        for column in "G":
            cell_name = "{}{}".format(column, row)
            if ws[cell_name].value < 6:
                rent_percentage = 0
            elif 6 <= ws[cell_name].value <= 7:
                rent_percentage = 0.25
            elif 8 <= ws[cell_name].value <= 16:
                rent_percentage = 0.5
            elif 17 <= ws[cell_name].value <= 24:
                rent_percentage = 0.75
            else:
                rent_percentage = 1
            print(str(ws[cell_name].value)+":"+str(rent_percentage))
        for column in "H":
            cell_name = "{}{}".format(column, row)
            ws[cell_name] = rent_percentage
    #Final save to confirm changes
    wb.save(filename=file_name_string)




file_entry = ttk.Entry(mainframe, width=20, textvariable=file_name)
file_entry.grid(column=2, row=1, sticky=(W,E))

month_entry = ttk.Entry(mainframe, width=2, textvariable=rent_month)
month_entry.grid(column=2, row=2, sticky=(W,E))

year_entry = ttk.Entry(mainframe, width=4, textvariable=rent_year)
year_entry.grid(column=2, row=3, sticky=(W,E))

ttk.Button(mainframe, text="Calculate Rent", command=calculate_rent).grid(column=3, row=4, sticky=W)
ttk.Label(mainframe, textvariable=status).grid(column=1, row=4, sticky=(W,E))
ttk.Label(mainframe, text="Please enter the filename").grid(column=1,row=1, sticky=W)
ttk.Label(mainframe, text="Please enter the month").grid(column=1,row=2, sticky=W)
ttk.Label(mainframe, text="Please enter the year").grid(column=1,row=3, sticky=W)

for child in mainframe.winfo_children(): child.grid_configure(padx=5, pady=5)
file_entry.focus()
root.bind('<Return>', calculate_rent)

root.mainloop()
